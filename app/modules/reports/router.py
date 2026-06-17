import csv
import io
from fastapi import APIRouter, Depends, Request, Query
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.database import get_db
from app.models.license import License
from app.models.finance import Finance
from app.models.exporter import Exporter
from app.models.product import Product
from app.models.market import Market
from app.modules.auth.deps import require_role
from app.templates import render

router = APIRouter(prefix="/reports", tags=["reports"],
                   dependencies=[Depends(require_role("admin", "developer", "employee"))])


@router.get("")
async def reports_page(request: Request, db: Session = Depends(get_db)):
    total_exporters = db.query(func.count(Exporter.id)).scalar()
    total_products = db.query(func.count(Product.id)).scalar()
    total_licenses = db.query(func.count(License.id)).scalar()
    pending_licenses = db.query(func.count(License.id)).where(License.status == "pending").scalar()
    approved_licenses = db.query(func.count(License.id)).where(License.status == "approved").scalar()
    total_fees = db.query(func.coalesce(func.sum(Finance.amount), 0)).scalar()
    paid_fees = db.query(func.coalesce(func.sum(Finance.amount), 0)).where(Finance.status == "paid").scalar()

    return render("reports/index.html", request=request, show_nav=True,
                  total_exporters=total_exporters,
                  total_products=total_products,
                  total_licenses=total_licenses,
                  pending_licenses=pending_licenses,
                  approved_licenses=approved_licenses,
                  total_fees=total_fees,
                  paid_fees=paid_fees)


@router.get("/export/csv")
async def export_csv(
    entity: str = Query("licenses"),
    db: Session = Depends(get_db),
):
    output = io.StringIO()
    writer = csv.writer(output)

    if entity == "licenses":
        writer.writerow(["ID", "المنتج", "المصدر", "السوق", "الحالة", "تاريخ الإنشاء", "تاريخ الاعتماد"])
        rows = db.execute(
            select(License).order_by(License.id.desc())
        ).scalars().all()
        for r in rows:
            writer.writerow([r.id, r.product_id, r.exporter_id, r.market_id, r.status,
                           r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                           r.approved_at.strftime("%Y-%m-%d") if r.approved_at else ""])

    elif entity == "finance":
        writer.writerow(["ID", "الترخيص", "المصدر", "المبلغ", "نوع الرسوم", "الحالة", "تاريخ الإنشاء", "تاريخ الدفع"])
        rows = db.execute(
            select(Finance).order_by(Finance.id.desc())
        ).scalars().all()
        for r in rows:
            writer.writerow([r.id, r.license_id, r.exporter_id, r.amount, r.fee_type, r.status,
                           r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                           r.paid_at.strftime("%Y-%m-%d") if r.paid_at else ""])

    elif entity == "exporters":
        writer.writerow(["ID", "الشركة", "المالك", "البريد", "الهاتف", "السجل التجاري", "العنوان"])
        rows = db.execute(select(Exporter).order_by(Exporter.id.desc())).scalars().all()
        for r in rows:
            writer.writerow([r.id, r.company_name, r.owner_name, r.email, r.phone, r.commercial_registry, r.address])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={entity}.csv"},
    )


@router.get("/export/xlsx")
async def export_xlsx(
    entity: str = Query("licenses"),
    db: Session = Depends(get_db),
):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="0D5E2E", end_color="0D5E2E", fill_type="solid")

    if entity == "licenses":
        headers = ["ID", "المنتج", "المصدر", "السوق", "الحالة", "تاريخ الإنشاء", "تاريخ الاعتماد"]
        ws.append(headers)
        rows = db.execute(select(License).order_by(License.id.desc())).scalars().all()
        for r in rows:
            ws.append([r.id, r.product_id, r.exporter_id, r.market_id, r.status,
                       r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                       r.approved_at.strftime("%Y-%m-%d") if r.approved_at else ""])

    elif entity == "finance":
        headers = ["ID", "الترخيص", "المصدر", "المبلغ", "نوع الرسوم", "الحالة", "تاريخ الإنشاء", "تاريخ الدفع"]
        ws.append(headers)
        rows = db.execute(select(Finance).order_by(Finance.id.desc())).scalars().all()
        for r in rows:
            ws.append([r.id, r.license_id, r.exporter_id, r.amount, r.fee_type, r.status,
                       r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                       r.paid_at.strftime("%Y-%m-%d") if r.paid_at else ""])

    elif entity == "exporters":
        headers = ["ID", "الشركة", "المالك", "البريد", "الهاتف", "السجل التجاري", "العنوان"]
        ws.append(headers)
        rows = db.execute(select(Exporter).order_by(Exporter.id.desc())).scalars().all()
        for r in rows:
            ws.append([r.id, r.company_name, r.owner_name, r.email, r.phone, r.commercial_registry, r.address])

    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity}.xlsx"},
    )


@router.get("/view")
async def entity_view(request: Request, entity: str = Query(...), db: Session = Depends(get_db)):
    models = {"licenses": License, "finance": Finance, "exporters": Exporter}
    labels = {"licenses": "التراخيص", "finance": "المالية", "exporters": "المصدرين"}
    icons = {"licenses": "bi-file-text", "finance": "bi-currency-exchange", "exporters": "bi-people"}
    if entity not in models:
        from app.core.exceptions import NotFound
        raise NotFound("نوع التقرير غير موجود")
    model = models[entity]
    records = db.execute(select(model).order_by(model.id.desc())).scalars().all()
    return render("reports/entity.html", request=request, entity=entity, records=records,
                  label=labels[entity], icon=icons[entity], show_nav=True)


@router.get("/export/one/csv")
async def export_single_csv(
    entity: str = Query(...), record_id: int = Query(...), db: Session = Depends(get_db)
):
    models = {"licenses": License, "finance": Finance, "exporters": Exporter}
    if entity not in models:
        from app.core.exceptions import NotFound
        raise NotFound("نوع التقرير غير موجود")
    model = models[entity]
    r = db.execute(select(model).where(model.id == record_id)).scalar_one_or_none()
    if not r:
        from app.core.exceptions import NotFound
        raise NotFound("السجل غير موجود")
    output = io.StringIO()
    writer = csv.writer(output)
    if entity == "licenses":
        writer.writerow(["ID", "المنتج", "المصدر", "السوق", "الحالة", "تاريخ الإنشاء", "تاريخ الاعتماد"])
        writer.writerow([r.id, r.product_id, r.exporter_id, r.market_id, r.status,
                         r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                         r.approved_at.strftime("%Y-%m-%d") if r.approved_at else ""])
    elif entity == "finance":
        writer.writerow(["ID", "الترخيص", "المصدر", "المبلغ", "نوع الرسوم", "الحالة", "تاريخ الإنشاء", "تاريخ الدفع"])
        writer.writerow([r.id, r.license_id, r.exporter_id, r.amount, r.fee_type, r.status,
                         r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                         r.paid_at.strftime("%Y-%m-%d") if r.paid_at else ""])
    elif entity == "exporters":
        writer.writerow(["ID", "الشركة", "المالك", "البريد", "الهاتف", "السجل التجاري", "العنوان"])
        writer.writerow([r.id, r.company_name, r.owner_name, r.email, r.phone, r.commercial_registry, r.address])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={entity}_{record_id}.csv"},
    )


@router.get("/export/one/xlsx")
async def export_single_xlsx(
    entity: str = Query(...), record_id: int = Query(...), db: Session = Depends(get_db)
):
    import openpyxl
    models = {"licenses": License, "finance": Finance, "exporters": Exporter}
    if entity not in models:
        from app.core.exceptions import NotFound
        raise NotFound("نوع التقرير غير موجود")
    model = models[entity]
    r = db.execute(select(model).where(model.id == record_id)).scalar_one_or_none()
    if not r:
        from app.core.exceptions import NotFound
        raise NotFound("السجل غير موجود")
    wb = openpyxl.Workbook()
    ws = wb.active
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="0D5E2E", end_color="0D5E2E", fill_type="solid")
    if entity == "licenses":
        ws.append(["ID", "المنتج", "المصدر", "السوق", "الحالة", "تاريخ الإنشاء", "تاريخ الاعتماد"])
        ws.append([r.id, r.product_id, r.exporter_id, r.market_id, r.status,
                   r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                   r.approved_at.strftime("%Y-%m-%d") if r.approved_at else ""])
    elif entity == "finance":
        ws.append(["ID", "الترخيص", "المصدر", "المبلغ", "نوع الرسوم", "الحالة", "تاريخ الإنشاء", "تاريخ الدفع"])
        ws.append([r.id, r.license_id, r.exporter_id, r.amount, r.fee_type, r.status,
                   r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
                   r.paid_at.strftime("%Y-%m-%d") if r.paid_at else ""])
    elif entity == "exporters":
        ws.append(["ID", "الشركة", "المالك", "البريد", "الهاتف", "السجل التجاري", "العنوان"])
        ws.append([r.id, r.company_name, r.owner_name, r.email, r.phone, r.commercial_registry, r.address])
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity}_{record_id}.xlsx"},
    )
